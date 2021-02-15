################################################################################################
# views.py: Part of UploadFile Package, Used for Upload File Functionality
# Classes:
# Upload
#
################################################################################################

from django.shortcuts import render
from django.http.response import JsonResponse
from django.core.files.base import ContentFile
from django.core.serializers.json import DjangoJSONEncoder
from django.apps import apps	

from rest_framework.request import HttpRequest, Request
from rest_framework.views import APIView
from rest_framework import status

from common_files.read_configuration import read_config	
from common_files.jwt_token import make_jwt_token, extract_jwt_info, verify_jwt_token
from common_files.read_logger import get_logger
from common_files.create_response import create_failure_modified, create_success_modified, create_failure, create_success

from .serializers import tblmdatabasesettingsSerializer, sessionmainSerializer
from dashboard.models import TblMDatabasesettings, Sessionmain

import requests,json,csv,os,io,base64,hashlib,urllib
import pandas as pd
from openpyxl import load_workbook

########################################################################################################
# Class Upload - Collection of functions for reading and dumping Excel/CSV to the server
# Functions:
# post - is used to take input from the client and send response to the client.
# save_file - is used to save File(excel,csv) on machine using os module.
# read_csv_file - is used to read csv file from request and process it.
# read_excel_file - is used to read excel file from request and process it.
# excel_to_json - convert records of excel into json 
# verified_md5 - fetch md5 checksum from the request and then compare it from internal created checksum
########################################################################################################

class Upload(APIView):
    def __init__(self):
        self.logger = get_logger()
        self.config = read_config()

    def post(self, request : Request) -> JsonResponse:
        try:
            token = verify_jwt_token(request.META['HTTP_AUTHORIZATION'])
            if(token == 0):
                self.logger.error("Error in Upload File Import API token: ")
                return JsonResponse(create_failure(400, 'Invalid token', 'Fail'))
        except Exception as e:
            self.logger.error("Error in Authentication."+str(e))
            return JsonResponse(create_failure_modified(500, 'Authentication Error', 'Failed', '', 'failed'))
        try:
            self.uniqueID = request.data['uniqueID']
        except Exception as e:
            self.logger.error("Please provide uniqueid "+str(e))
            return JsonResponse(create_failure_modified(500, 'Please provide uniqueid', 'Failed', '', 'failed'))
        try:
            self.fileName = request.data['fileName'] 
        except Exception as e:
            self.logger.error("Error in FileName "+str(e))
            return JsonResponse(create_failure_modified(500, 'Error in FileName ', 'Failed', '', 'failed'))
        try:  
            if(not request.data['file']):
                return JsonResponse(create_failure_modified(500, 'Failed to Upload File', 'Failed', '', 'failed'))
            elif(len(self.uniqueID.strip()) == 0):
                return JsonResponse(create_failure_modified(500, 'Failed to Upload File', 'Failed', '', 'failed'))
            else:
                if(self.save_file(request) != True):
                    return JsonResponse(create_failure_modified(500, 'File is Corrupted', 'Failed', '', 'failed'))
                if(self.fileName.split('.')[-1] == "csv"):
                    response = self.read_csv_file(request)
                elif(self.fileName.split('.')[-1] == 'xlsx'):
                    response = self.read_excel_file(request)
                else:
                    JsonResponse(create_failure_modified(500, 'Invalid file type , Please upload Excel or CSV file', 'Failed', '', 'failed'))       
        except Exception as e:
            self.logger.error("Error in API"+str(e))
            return JsonResponse(create_failure_modified(500, 'API Fail due to Exception', 'Failed', '', 'failed'))
        return JsonResponse(response)

    def save_file(self, request):
        try: 
            directory = self.config['ENV_VARIABLE']['file_path'] + str(self.uniqueID)+'/'
            if not os.path.exists(directory):
                os.makedirs(directory)
            open(directory + self.fileName, 'wb').write(request.data['file'].read())
            if(not self.verified_md5(request)):
                return False
            else:
                return True
        except Exception as e:
            self.logger.error("Occured error while saving file : "+str(e))
            return False

    def read_csv_file(self, request):
        try:
            textQual = request.data['textqualifier']
            if(len(textQual) == 0):
                textQual = ' '
        except Exception as e:
            self.logger.error("Text Qualifier is not proper - "+str(e))
            return(create_failure_modified(500, 'Text Qualifier is not proper', 'Failed', '', 'failed'))
        try:
            rowDel = base64.b64decode(request.data['rowdelimeter']).decode("utf-8")
        except Exception as e:
            self.logger.error("Row Delimeter is not proper - "+str(e))
            return(create_failure_modified(500, 'Row Delimeter is not proper', 'Failed', '', 'failed'))
        try:
            colDel = base64.b64decode(request.data['columndelimeter']).decode("utf-8")
        except Exception as e:
            self.logger.error("Column Delimeter is not proper - "+str(e))
            return(create_failure_modified(500, 'Column Delimeter is not proper', 'Failed', '', 'failed'))
        try:
            noRowToSkip = int(request.data['numberRowToSkip'])
        except Exception as e:
            self.logger.error("Row To Skip is not proper - "+str(e))
            return(create_failure_modified(500, 'Row to Skip is not proper', 'Failed', '', 'failed'))
        try:
            firstRowHeader = request.data['iscolumnnameinfirstrow']
            if(firstRowHeader == 'true'):
                firstRowHeader = 0
            else:
                firstRowHeader = None
        except Exception as e:
            self.logger.error("Is Column Name in First Row is not proper - "+str(e))
            return(create_failure_modified(500, 'Is Column Name in First Row not proper', 'Failed', '', 'failed'))            
        try:
            directory = self.config['ENV_VARIABLE']['file_path'] + str(self.uniqueID) + '/' + self.fileName
            with open(directory) as csv_data:
                data = csv_data.read()
            if(textQual != " "):
                data = data.replace(textQual,"'")
        except Exception as e:
            self.logger.error("Error in opening directory - "+str(e))
            return(create_failure_modified(500, 'Error in opening directory', 'Failed', '', 'failed')) 
        try:
            if(rowDel == '/r/n' or rowDel == '//r//n' or rowDel == '\\r\\n' or rowDel == '\r\n'):
                data = pd.read_csv(io.StringIO(data),sep=colDel,header=firstRowHeader,skiprows=noRowToSkip,quotechar="'")
            elif(rowDel == '/n'):
                data = pd.read_csv(io.StringIO(data),sep=colDel,lineterminator='\n',header=firstRowHeader,skiprows=noRowToSkip,quotechar="'")
            elif(rowDel == '/r'):
                data = pd.read_csv(io.StringIO(data),sep=colDel,lineterminator='\r',header=firstRowHeader,skiprows=noRowToSkip,quotechar="'")
            else:
                data = pd.read_csv(io.StringIO(data),sep=colDel,lineterminator=rowDel,header=firstRowHeader,skiprows=noRowToSkip,quotechar="'")
            dataNa = data.fillna('')
            dataFirstFiftyRecords = dataNa.head(50)
            jsonString = dataFirstFiftyRecords.to_json(orient="records")
            jsonData = json.loads(jsonString)
            if(len(jsonData) > 0 ):
                response = create_success_modified('File uploaded Successfully', jsonData, 'uploaded')
            else:
                response = create_failure_modified(500, 'Invalid Delimeter Provided', 'Failed', '', 'failed') 
        except Exception as e:
            self.logger.error("Error in CSV File- "+str(e))
            return(create_failure_modified(500, 'Error in CSV File', 'Failed', '', 'failed'))
        return(response)

    def read_excel_file(self, request):
        try:        
            directory = self.config['ENV_VARIABLE']['file_path'] + str(self.uniqueID) + '/' + self.fileName
        except Exception as e:
            self.logger.error("Error in Request Parameters- "+str(e))
            return(create_failure_modified(500, 'Error in Request Parameters', 'Failed', '', 'failed'))  
        try:
            workBook = load_workbook(directory)
            workSheets = list(workBook.sheetnames)
            jsonData = list()
            for sheet in workSheets:
                try:
                    workSheetVar=workBook[sheet]
                    jsonChild = self.excel_to_json(workSheetVar)
                    if(jsonChild == False):
                        return(create_failure_modified(500, 'Error in Excel File', 'Failed', '', 'failed'))  
                    jsonObject = {}
                    jsonObject["Sheet"] = sheet
                    jsonObject['Data'] = jsonChild['Data']
                    jsonObject['ExcelDataValue'] = jsonChild['ExcelDataValue']
                    jsonObject['EmptyFlag'] = jsonChild['EmptyFlag']
                    jsonData.append(jsonObject)
                except Exception as e:
                    self.logger.error("Error in Reterving Data from excel file- "+str(e))
                    return(create_failure_modified(500, 'Error in Reterving Data from excel file', 'Failed', '', 'failed'))
        except Exception as e:
            self.logger.error("Error in Reterving Data- "+str(e))
            return(create_failure_modified(500, 'Error in Reterving Data', 'Failed', '', 'failed'))
        try:            
            if(len(jsonData) > 0 ):
                response = create_success_modified('File uploaded Successfully', jsonData, 'uploaded')
            else:
                response = create_failure_modified(500, 'Invalid Delimeter Provided', 'Failed', '', 'failed')
        except Exception as e:
            self.logger.error("Error in Excel File- "+str(e))
            return(create_failure_modified(500, 'Error in Excel File', 'Failed', '', 'failed'))
        return(response)

    def excel_to_json(self,excel_sheet):
        try:
            data = excel_sheet.values
            cols = next(data)[:]
            dataVar = list(data)
            dataFrame = pd.DataFrame(dataVar, columns=cols)
            dataFrmaeNa = dataFrame.fillna('')
            headers = list(dataFrmaeNa.columns.values)
            dataFirstFiftyRecords = dataFrmaeNa.head(50)
            jsonString = dataFirstFiftyRecords.to_dict(orient="records")
            jsonString = json.dumps(jsonString, sort_keys=True, indent=1, cls=DjangoJSONEncoder)
            jsonData = {}
            jsonData['Data'] = json.loads(jsonString)
            excelDataValue = list()
            nullList = list()
            for header in headers:
                try:
                    dataObj = {}
                    dataObj['HeaderName'] = header
                    dataObj['HeaderValue'] = list(dataFrmaeNa[header])[:10]
                    result = all(x == '' for x in dataObj['HeaderValue'])
                    nullList.append(result)
                    excelDataValue.append(dataObj)
                except Exception as e:
                    self.logger.error("Error in Excel File in Excel2JSON- "+str(e))
                    return False
            jsonData['ExcelDataValue'] = excelDataValue
            jsonData['EmptyFlag'] = all(x for x in nullList)
        except Exception as e:
            self.logger.error("Error in Excel File in Excel2JSON- "+str(e))
            return False
        return(jsonData)

    def verified_md5(self,request):
        try:        
            md5Checksum = request.data['md5HashValue']
            directory = self.config['ENV_VARIABLE']['file_path'] + str(self.uniqueID) +'/' + self.fileName
            md5Hash = hashlib.md5()
            with open(directory,'rb') as dataFile:
                fileContent = dataFile.read()
            md5Hash.update(fileContent)
            digest = md5Hash.hexdigest()        
            if(digest != md5Checksum):
                return True
        except Exception as e:
            self.logger.error("Occured error while verifying md5 checksum : "+str(e))
            return False
        self.logger.info("MDF Checksum matched successfully")
        return True

########################################################################################################
# Class FetchDatabaseName - Collection of functions for fetching destingation database Name 
# Functions:
# fetch - is used to take input from the client and then return database name to the client for corresponding todolist id
########################################################################################################
class FetchDatabaseName():
    def __init__(self):
        self.logger = get_logger()

    def fetch(self,sessionId):
        try:
            databaseName = ''
            sessionMainObject = Sessionmain.objects.filter(sessionmainid=sessionId)
            sessionMainData = sessionmainSerializer(sessionMainObject, many=True)
            if sessionMainObject.exists():
                try:
                    for sessionMainRecord in sessionMainData.data:
                        try:
                            dbidValue = sessionMainRecord['dbid']
                            tblmdatabasesettingsObject = TblMDatabasesettings.objects.filter(databasesettingsid=dbidValue)
                            tblmdatabasesettingsData= tblmdatabasesettingsSerializer(tblmdatabasesettingsObject, many=True)
                            if tblmdatabasesettingsObject.exists():
                                for dbidRecord in tblmdatabasesettingsData.data:
                                    databaseName = dbidRecord['databasename']                  
                        except Exception as e:
                            self.logger.error("Error in TblMDatabaseSettings Table: "+str(e))
                except Exception as e:
                    self.logger.error("Error in SessionMain Table: "+str(e))
        except Exception as e:
            self.logger.error("Error in Fetching Data."+str(e))
        return databaseName